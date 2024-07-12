import tkinter as tk
import os
from tkinter import ttk
from tkinter import filedialog
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import sys

# 定義標題和內容的格式
header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
header_font = Font(name="Calibri", bold=False, color='FFFFFF', sz=11)
content_font = Font(name="Calibri", bold=False, color='000000', sz=11)


def choose_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        file_path_entry.delete(0, tk.END)
        file_path_entry.insert(0, file_path)
        load_sheets(file_path)


def load_sheets(file_path):
    xl = pd.ExcelFile(file_path)
    sheets = xl.sheet_names
    sheet_combobox['values'] = sheets
    if sheets:
        sheet_combobox.current(0)
        sheet_combobox.bind("<<ComboboxSelected>>", lambda event: load_columns(file_path, sheet_combobox.get()))


def load_columns(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    df.columns = [col.replace('\n', ' ') for col in df.columns]
    columns = df.columns.tolist()
    column_combobox['values'] = columns
    filter_column_combobox['values'] = columns
    if len(columns) > 0:
        column_combobox.current(0)
    
    column_listbox.delete(0, tk.END)
    for col in columns:
        column_listbox.insert(tk.END, col)
    
    select_all_var.set(False)


def update_filter_values(*args):
    selected_column = filter_column_combobox.get()
    file_path = file_path_entry.get()
    selected_sheet = sheet_combobox.get()
    df = pd.read_excel(file_path, sheet_name=selected_sheet)
    df.columns = [col.replace('\n', ' ') for col in df.columns]
    if selected_column in df.columns:
        unique_values = sorted([str(value) for value in df[selected_column].unique()])
        filter_value_combobox['values'] = unique_values
    else:
        filter_value_combobox['values'] = []


def get_selected_columns():
    selected_indices = column_listbox.curselection()
    selected_columns = [column_listbox.get(i) for i in selected_indices]
    return selected_columns


def split_file():
    file_path = file_path_entry.get()
    selected_sheet = sheet_combobox.get()
    selected_column = column_combobox.get()
    filter_column = filter_column_combobox.get()
    filter_value = filter_value_combobox.get()
    df = pd.read_excel(file_path, sheet_name=selected_sheet)
    df.columns = [col.replace('\n', ' ') for col in df.columns]
    
    if filter_value and filter_column:
        df = df[df[filter_column] == filter_value]

    unique_values = df[selected_column].unique()
    print(unique_values)
    output_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.realpath(__file__))
    output_dir = os.path.join(output_dir, "output")
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for value in unique_values:
        df_filtered = df[df[selected_column] == value]
        selected_columns = get_selected_columns()
        if selected_columns:
            df_filtered = df_filtered[selected_columns]
        new_file_name = os.path.join(output_dir, f"{os.path.splitext(os.path.basename(file_path))[0]}_{value}.xlsx")
        df_filtered.to_excel(new_file_name, index=False)

        wb = openpyxl.load_workbook(new_file_name)
        ws = wb.active

        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
            for cell in col:
                cell.fill = header_fill
                cell.font = header_font

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = content_font
        
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = (length + 2) * 1.2
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

        wb.save(new_file_name)
        print(f"已創建檔案：{new_file_name}")


def select_all_columns():
    if select_all_var.get():
        column_listbox.select_set(0, tk.END)
    else:
        column_listbox.select_clear(0, tk.END)


if __name__ == "__main__":
    root = tk.Tk()
    root.title("檔案分割工具")

    # 創建一個標籤來顯示檔案路徑輸入框旁的紅色星號
    required_label = tk.Label(root, text='*', fg='red')
    required_label.grid(row=0, column=0, sticky='w')


    tk.Label(root, text="選擇檔案：").grid(row=0, column=0, padx=10, pady=10)
    file_path_entry = tk.Entry(root, width=50)
    file_path_entry.grid(row=0, column=1, padx=10, pady=10)
    choose_file_button = tk.Button(root, text="瀏覽", command=choose_file)
    choose_file_button.grid(row=0, column=2, padx=10, pady=10)

    required_label = tk.Label(root, text='*', fg='red')
    required_label.grid(row=1, column=0, sticky='w')
    tk.Label(root, text="選擇工作表：").grid(row=1, column=0, padx=10, pady=10)
    sheet_combobox = ttk.Combobox(root, width=47)
    sheet_combobox.grid(row=1, column=1, padx=10, pady=10)

    required_label = tk.Label(root, text='*', fg='red')
    required_label.grid(row=2, column=0, sticky='w')
    tk.Label(root, text="選擇列：").grid(row=2, column=0, padx=10, pady=10)
    column_combobox = ttk.Combobox(root, width=47)
    column_combobox.grid(row=2, column=1, padx=10, pady=10)

    tk.Label(root, text="選擇過濾列：").grid(row=3, column=0, padx=10, pady=10)
    filter_column_combobox = ttk.Combobox(root, width=47)
    filter_column_combobox.grid(row=3, column=1, padx=10, pady=10)
    filter_column_combobox.bind('<<ComboboxSelected>>', update_filter_values)

    tk.Label(root, text="選擇過濾值：").grid(row=4, column=0, padx=10, pady=10)
    filter_value_combobox = ttk.Combobox(root, width=47)
    filter_value_combobox.grid(row=4, column=1, padx=10, pady=10)

    required_label = tk.Label(root, text='*', fg='red')
    required_label.grid(row=5, column=0, sticky='w')
    tk.Label(root, text="選擇輸出欄位：").grid(row=5, column=0, padx=10, pady=10)
    select_all_var = tk.BooleanVar()
    select_all_checkbutton = tk.Checkbutton(root, text="全選", variable=select_all_var, command=select_all_columns)
    select_all_checkbutton.grid(row=5, column=1, sticky='w', padx=10, pady=10)

    column_listbox = tk.Listbox(root, selectmode=tk.MULTIPLE, width=47, height=10)
    column_listbox.grid(row=6, column=1, padx=10, pady=10)

    split_button = tk.Button(root, text="分割檔案", command=split_file)
    split_button.grid(row=7, column=0, columnspan=2, padx=10, pady=10)

    root.mainloop()
